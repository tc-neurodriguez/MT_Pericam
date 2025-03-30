import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the workbook
input_file = r"C:\Users\t_rod\Box\SY5Y transduction Images\Experiment 3_9_25\MT-Pericam_3_9_25_FCCP.xlsx"  # Replace with your file path
output_file = r"C:\Users\t_rod\Box\SY5Y transduction Images\Experiment 3_9_25\MT-Pericam_3_9_25_FCCP_python.xlsx"

try:
    wb = load_workbook(input_file)
    ws = wb.active
    print(f"Successfully loaded workbook: {input_file}")
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit()

wavelength_data = {}  # Stores organized data by wavelength
current_wavelength = None
current_headers = []
data_rows = []

# Convert rows to a controllable iterator
rows = iter(ws.iter_rows(values_only=True))

for row in rows:
    if row[0] and 'Read' in str(row[0]):
        print(f"Found new wavelength: {row[0]}")

        # Save previous data
        if current_wavelength and data_rows:
            df = pd.DataFrame(data_rows, columns=current_headers)
            df['Time'] = pd.to_timedelta(df['Time'])
            if current_wavelength in wavelength_data:
                wavelength_data[current_wavelength] = pd.concat([wavelength_data[current_wavelength], df])
            else:
                wavelength_data[current_wavelength] = df
            print(f"Saved data for wavelength: {current_wavelength}")
            data_rows = []

        # Extract wavelength and get headers
        current_wavelength = row[0].split(':')[-1].strip()
        try:
            header_row = next(rows)  # Get next row after "Read"
            print(f"Headers for {current_wavelength}: {header_row}")
        except StopIteration:
            print("Reached end of file while reading headers.")
            break

        # Standardize headers: Time, Temperature + well IDs
        current_headers = ['Time', 'Temperature']
        current_headers += [cell for cell in header_row[2:] if cell is not None]
        print(f"Processed headers: {current_headers}")

    elif current_wavelength and row[0]:
        # Capture only rows with actual data
        if any(cell is not None for cell in row[2:len(current_headers)]):
            data_row = [
                row[0],  # Time
                row[1],  # Temperature
                *row[2:2 + len(current_headers) - 2]  # Well data
            ]
            data_rows.append(data_row)
            print(f"Added data row: {data_row}")

# Process final wavelength
if current_wavelength and data_rows:
    df = pd.DataFrame(data_rows, columns=current_headers)
    df['Time'] = pd.to_timedelta(df['Time'])
    if current_wavelength in wavelength_data:
        wavelength_data[current_wavelength] = pd.concat([wavelength_data[current_wavelength], df])
    else:
        wavelength_data[current_wavelength] = df
    print(f"Final data saved for wavelength: {current_wavelength}")

# Create sorted sheets for each wavelength
if wavelength_data:
    for wavelength, df in wavelength_data.items():
        df = df.sort_values('Time').reset_index(drop=True)
        if wavelength in wb.sheetnames:
            del wb[wavelength]
        new_ws = wb.create_sheet(wavelength)
        for r in dataframe_to_rows(df, index=False, header=True):
            new_ws.append(r)
        print(f"Created sheet for wavelength: {wavelength}")

    # Save the modified workbook
    wb.save(output_file)
    print(f"Workbook saved as: {output_file}")
else:
    print("No data was processed. Check the input file format.")